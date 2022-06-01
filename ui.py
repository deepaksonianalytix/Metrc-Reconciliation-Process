from datetime import timedelta, date
import datetime
import PySimpleGUI as sg
import threading
import queue
import ctypes
import os
from openpyxl import load_workbook
from dateutil.parser import parse
from MetrcDownload import RunMetrc

sg.theme('DarkBlue3')
gui_queue = queue.Queue()


def mBox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)

def download_report(start_date, end_date, username, password, email, report_data):
    run_metrc = RunMetrc()
    run_metrc.gui_queue = gui_queue
    run_status = run_metrc.run(start_date, end_date, username, password, email, report_data)
    return run_status



def load_setting():
    settingFile = 'MetrcSettingSheet.xlsx'
    if not os.path.exists(settingFile):
        mBox('Error', f'Setting Sheet not found in Macro Folder or you Name rename Setting File. '
                      f'\nSetting Sheet name must be "{settingFile}"', 0)
        return False
    else:
        load_file = load_workbook(settingFile)
        setting_sheets = load_file.sheetnames
        cred_sheet = '--CREDS--'
        report_sheet = '--REPORTS--'
        if cred_sheet not in setting_sheets or report_sheet not in setting_sheets:
            mBox('Error', f'Vendor Credentials sheet or Report Sheet not found in Setting Sheet. '
                          f'\nCredentials sheet must be named with "{cred_sheet} and Report sheet must be names with {report_sheet}"', 0)
            return False
        else:
            creds = load_file[cred_sheet].values
            report = load_file[report_sheet].values
            creds_data = [list(items) for items in creds if items]
            report_data = [list(items) for items in report if items]
            if not creds_data or not report_data:
                mBox('Error', f'No Data found in Creds Sheet. or Report sheet', 0)
                return False
            else:
                return creds_data, report_data

def run_gui(thread=None):
    version = f'MRP_{date.today()}'
    end_date = (date.today() - timedelta(days=1)).strftime('%m/%d/%Y')
    start_date = datetime.datetime.strptime(end_date, '%m/%d/%Y').replace(day=1).strftime('%m/%d/%Y')
    creds_data, report_data = load_setting()

    layout = [
        [
            sg.Text('Metrc Reconciliation Process', size=(35, 1), font=('Corbel', 20), justification='center')
        ],
        [
            sg.CalendarButton('Start Date', size=(12, 1), format='%m/%d/%Y', key='start_date_btn', enable_events=True),
            sg.Input(start_date, size=(12, 1), font=('Corbel', 11), key='start_date', disabled=True,
                     justification='center', enable_events=True, readonly=True),
        ],
        [
            sg.CalendarButton('End Date', size=(12, 1), format='%m/%d/%Y', key='end_date_btn', enable_events=True),
            sg.Input(end_date, size=(12, 1), font=('Corbel', 11), key='end_date', disabled=True,
                     justification='center', enable_events=True, readonly=True),
        ],
        [
            sg.Text("Choose folder: "), sg.Input(),
            sg.FolderBrowse(initial_folder=os.getcwd(), key="filepath"),
        ],
        [
            sg.OK('Download Reports', key='download', size=(18, 1), font=('Corbel', 10), pad=((5, 5), (10, 0))),
            sg.OK('Reconcile', key='reconcile', size=(18, 1), font=('Corbel', 10), pad=((5, 5), (10, 0))),
            sg.Exit('Exit', key='exit', size=(18, 1), font=('Corbel', 10), pad=((5, 5), (10, 0))),
        ],
        [
            sg.Text('Status :', size=(15, 1), justification='left', font=('Corbel', 11)),
        ],
        [
            sg.Multiline(size=(65, 7), font='courier 10', background_color='white', text_color='black', key='status',
                         autoscroll=True, enable_events=True, change_submits=False)
        ],
    ]

    main_window = sg.Window(version,
                            element_justification='left',
                            text_justification='left',
                            auto_size_text=True, resizable=True).Layout(layout).Finalize()
    main_window.Maximize()

    while True:
        event, values = main_window.Read(timeout=1000)
        main_window.refresh()


        if event in ('exit', None) or event == sg.WIN_CLOSED:
            main_window.close()
            break
        elif event == 'download':
            start_date = values['start_date']
            end_date = values['end_date']
            for creds in creds_data[1:]:
                username = creds[0]
                password = creds[1]
                email = creds[2]
            if parse(str(start_date)).date() > parse(str(end_date)).date():
                msg = f'\nError : Start Date should not be grater than End Date.'
                main_window['status'].Update(msg)
                continue
            main_window['download'].Update(disabled=True)
            main_window['exit'].Update(disabled=True)
            thread = threading.Thread(target=download_report,
                                      args=(start_date, end_date, username, password, email, report_data))
            thread.start()
        elif event == 'reconcile':
            file_path = values['filepath']
            if not file_path:
                msg = f'\nError : Path cannot be empty.'
                main_window['status'].Update(msg)


        if thread:
            if not thread.is_alive():
                main_window['download'].Update(disabled=False)
                main_window['exit'].Update(disabled=False)
                main_window.refresh()

        try:
            message = gui_queue.get_nowait()
        except:
            message = None
        if message:
            for key, value in message.items():
                if key == 'status':
                    main_window['status'].print(value)
                    main_window.refresh()
                if key == 'Success':
                    sg.Popup(value, title='Status')
            main_window.refresh()

if __name__ == '__main__':
    run_gui()